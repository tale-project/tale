"""XLSX text extraction using openpyxl.

Converts spreadsheet rows into plain text for indexing.
"""

import zipfile
from io import BytesIO

from loguru import logger
from openpyxl import load_workbook

MAX_UNCOMPRESSED_SIZE = 500 * 1024 * 1024  # 500 MB


async def extract_text_from_xlsx_bytes(
    xlsx_bytes: bytes,
    filename: str = "spreadsheet.xlsx",
) -> tuple[str, bool]:
    """Extract text from XLSX bytes.

    Args:
        xlsx_bytes: Raw XLSX bytes.
        filename: Filename for logging.

    Returns:
        Tuple of (extracted_text, vision_was_used). Vision is never used for XLSX.
    """
    logger.info(f"Processing XLSX: {filename}")

    try:
        zf = zipfile.ZipFile(BytesIO(xlsx_bytes))
        total = sum(info.file_size for info in zf.infolist())
        if total > MAX_UNCOMPRESSED_SIZE:
            raise ValueError(f"File exceeds maximum decompressed size ({total} bytes)")
        zf.close()
    except zipfile.BadZipFile:
        raise ValueError("Invalid or corrupt file")

    wb = load_workbook(BytesIO(xlsx_bytes), read_only=True, data_only=True)
    try:
        sheets_text: list[str] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows_text: list[str] = []

            for row in ws.iter_rows(values_only=True):
                if not any(cell is not None and str(cell).strip() for cell in row):
                    continue
                cells = [str(cell) if cell is not None else "" for cell in row]
                line = " | ".join(cells).strip()
                if line:
                    rows_text.append(line)

            if rows_text:
                sheet_header = f"--- Sheet: {sheet_name} ---"
                sheets_text.append(f"{sheet_header}\n" + "\n".join(rows_text))
    finally:
        wb.close()

    combined_text = "\n\n".join(sheets_text)
    logger.info(
        f"XLSX processing complete: {len(sheets_text)} sheets, {len(combined_text)} chars"
    )

    return combined_text, False
